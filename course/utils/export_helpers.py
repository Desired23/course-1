import csv
import io

from django.http import HttpResponse


def export_to_csv(headers, rows, filename):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def export_to_excel(headers, rows, filename, sheet_title='Sheet1'):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2563EB')
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        ws.append(row)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def export_workbook_to_excel(sheets, filename):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet in sheets:
        title = str(sheet.get('title') or 'Sheet')[:31]
        headers = sheet.get('headers') or []
        rows = sheet.get('rows') or []
        ws = wb.create_sheet(title=title)
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='2563EB')
            cell.alignment = Alignment(horizontal='center')

        for row in rows:
            ws.append(row)

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def export_sheets_to_zip_csv(sheets, filename):
    import zipfile

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as archive:
        for sheet in sheets:
            csv_buf = io.StringIO()
            writer = csv.writer(csv_buf)
            writer.writerow(sheet.get('headers') or [])
            writer.writerows(sheet.get('rows') or [])
            safe_name = ''.join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in str(sheet.get('title') or 'report'))
            archive.writestr(f'{safe_name}.csv', '\ufeff' + csv_buf.getvalue())
    buf.seek(0)

    response = HttpResponse(buf.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{filename}.zip"'
    return response
