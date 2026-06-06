import datetime
import io
import secrets

import openpyxl
from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_date
from rest_framework.exceptions import ValidationError

from instructors.models import Instructor
from subscription_plans.models import SubscriptionPlan, UserSubscription
from users.models import User


def parse_excel_bytes(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    return list(ws.iter_rows(min_row=2, values_only=True))


def _cell_date(value):
    if not value:
        return None
    if isinstance(value, datetime.datetime):
        return value if timezone.is_aware(value) else timezone.make_aware(value)
    if isinstance(value, datetime.date):
        return timezone.datetime(value.year, value.month, value.day, tzinfo=timezone.get_current_timezone())
    parsed = parse_date(str(value))
    if parsed:
        return timezone.datetime(parsed.year, parsed.month, parsed.day, tzinfo=timezone.get_current_timezone())
    return None


def import_subscription_plans(file_bytes, plan_id, admin_user):
    try:
        plan = SubscriptionPlan.objects.get(id=plan_id, is_deleted=False)
    except SubscriptionPlan.DoesNotExist:
        raise ValidationError(f'Subscription plan id={plan_id} does not exist.')

    rows = parse_excel_bytes(file_bytes)
    success = 0
    skipped = 0
    errors = []

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            email = str(row[0]).strip() if row and row[0] else ''
            if not email:
                continue

            start_date = _cell_date(row[1] if len(row) > 1 else None) or timezone.now()
            end_date = _cell_date(row[2] if len(row) > 2 else None)
            if end_date is None and plan.duration_days:
                end_date = start_date + datetime.timedelta(days=plan.duration_days)

            try:
                user = User.objects.get(email=email, is_deleted=False)
            except User.DoesNotExist:
                errors.append({'row': idx, 'email': email, 'reason': 'Email not found.'})
                continue

            exists = UserSubscription.objects.filter(
                user=user,
                plan=plan,
                status=UserSubscription.Status.ACTIVE,
                is_deleted=False,
            ).exists()
            if exists:
                skipped += 1
                errors.append({'row': idx, 'email': email, 'reason': 'Active subscription already exists for this plan.'})
                continue

            UserSubscription.objects.create(
                user=user,
                plan=plan,
                payment=None,
                status=UserSubscription.Status.ACTIVE,
                start_date=start_date,
                end_date=end_date,
            )
            success += 1

    return {'success': success, 'skipped': skipped, 'errors': errors}


def _unique_username(base):
    username = base[:255] or 'user'
    candidate = username
    counter = 1
    while User.objects.filter(username=candidate).exists():
        suffix = str(counter)
        candidate = f'{username[:255 - len(suffix)]}{suffix}'
        counter += 1
    return candidate


def import_users_bulk(file_bytes, admin_user):
    rows = parse_excel_bytes(file_bytes)
    created = 0
    updated = 0
    errors = []

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            email = str(row[0]).strip() if row and row[0] else ''
            full_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            username = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            role = str(row[3]).strip().lower() if len(row) > 3 and row[3] else 'student'
            password_raw = str(row[4]).strip() if len(row) > 4 and row[4] else ''

            if not email:
                continue
            if role not in ['student', 'instructor']:
                role = 'student'

            existing = User.objects.filter(email=email).first()
            if existing:
                update_fields = ['updated_at']
                if full_name:
                    existing.full_name = full_name
                    update_fields.append('full_name')
                existing.save(update_fields=update_fields)
                if role == 'instructor':
                    Instructor.objects.get_or_create(user=existing)
                updated += 1
                continue

            if not full_name:
                errors.append({'row': idx, 'email': email, 'reason': 'full_name is required for new users.'})
                continue

            if not username:
                username = _unique_username(email.split('@')[0])
            elif User.objects.filter(username=username).exists():
                username = _unique_username(username)

            if not password_raw:
                password_raw = secrets.token_urlsafe(12)

            user = User.objects.create(
                email=email,
                username=username,
                full_name=full_name,
                password_hash=make_password(password_raw),
                status=User.StatusChoices.ACTIVE,
            )
            if role == 'instructor':
                Instructor.objects.get_or_create(user=user)
            created += 1

    return {'created': created, 'updated': updated, 'errors': errors}


def import_course_grants(file_bytes, course_ids, admin_user):
    from courses.models import Course
    from enrollments.models import Enrollment
    from django.db.models import F

    courses = list(Course.objects.filter(id__in=course_ids, is_deleted=False, status=Course.Status.PUBLISHED))
    if not courses:
        raise ValidationError('No valid published courses found for the given IDs.')

    rows = parse_excel_bytes(file_bytes)
    success = 0
    skipped = 0
    errors = []

    OWNED = {Enrollment.Status.Active, Enrollment.Status.Complete, Enrollment.Status.SUSPENDED}

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            email = str(row[0]).strip() if row and row[0] else ''
            if not email:
                continue

            try:
                user = User.objects.get(email=email, is_deleted=False)
            except User.DoesNotExist:
                errors.append({'row': idx, 'email': email, 'reason': 'Email not found.'})
                continue

            for course in courses:
                existing = Enrollment.objects.filter(user=user, course=course).first()
                if existing:
                    if not existing.is_deleted and existing.status in OWNED:
                        skipped += 1
                        continue
                    existing.is_deleted = False
                    existing.deleted_at = None
                    existing.deleted_by = None
                    existing.status = Enrollment.Status.Active
                    existing.source = Enrollment.Source.GRANTED
                    if not existing.enrollment_date:
                        existing.enrollment_date = timezone.now()
                    existing.save(update_fields=['is_deleted', 'deleted_at', 'deleted_by', 'status', 'source', 'enrollment_date', 'updated_at'])
                else:
                    Enrollment.objects.create(
                        user=user,
                        course=course,
                        payment=None,
                        source=Enrollment.Source.GRANTED,
                        status=Enrollment.Status.Active,
                        enrollment_date=timezone.now(),
                    )
                    Course.objects.filter(id=course.id).update(total_students=F('total_students') + 1)
                success += 1

    return {'success': success, 'skipped': skipped, 'errors': errors}


def generate_course_grants_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Grant Courses'
    ws.append(['Email (*required)'])
    ws.append(['user@example.com'])
    _style_header(ws)
    return _workbook_bytes(wb)


def generate_subscription_import_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Import Subscriptions'
    ws.append(['Email (*required)', 'Start Date (YYYY-MM-DD)', 'End Date (YYYY-MM-DD)'])
    ws.append(['user@example.com', '2026-01-01', '2026-12-31'])
    _style_header(ws)
    return _workbook_bytes(wb)


def generate_users_import_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Import Users'
    ws.append(['Email (*required)', 'Full Name (*new users)', 'Username (optional)', 'User Type (student/instructor)', 'Password (optional)'])
    ws.append(['user@example.com', 'Nguyen Van A', 'nguyenvana', 'student', ''])
    _style_header(ws)
    return _workbook_bytes(wb)


def _style_header(ws):
    from openpyxl.styles import Font, PatternFill

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2563EB')


def _workbook_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
